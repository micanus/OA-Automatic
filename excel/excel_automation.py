import xlrd #xls
import openpyxl #xlsx only
from openpyxl import Workbook,load_workbook
import xlwings as xw #xlsx only

#파일 생성
workbook=Workbook()
worksheet=workbook.active #workbook 생성 시 기본 워크시트 바인딩
worksheet["A1"]="엑셀 자동화" #cell input
workbook.save("start.xlsx")

#파일 열기
wb=load_workbook("start.xlsx")
print(type(wb)) #<class 'openpyxl.workbook.workbook.Workbook'>
print(wb.sheetnames)

#sheet
sheetname=wb.sheetnames[0] #workbook sheet list의 첫번째 원소
ws=wb[sheetname]
ws=wb.active #첫 번째 엑셀시트에 바로 접근 #첫 번째 엑셀시트=활성시트
print(type(ws)) #<class 'openpyxl.worksheet.worksheet.Worksheet'>